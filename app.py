import streamlit as st
import pandas as pd
from datetime import datetime
from io import BytesIO

# Configuración de librerías para diseño de Excel
try:
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_STYLING = True
except ImportError:
    EXCEL_STYLING = False

st.set_page_config(page_title="Gestión de Repuestos - Reporte Maestro", layout="wide")

# --- BANNER ---
hoy = datetime.now()
st.markdown(f"""
    <div style="background: linear-gradient(90deg, #1F4E78 0%, #2E75B6 100%); padding: 20px; border-radius: 15px; color: white; text-align: center; margin-bottom: 20px;">
        <h1>🛠️ CONTROL MAESTRO DE REPUESTOS</h1>
        <p>Prioridad Centros <b>GO</b> | Detalle de Producto y Serie | {hoy.strftime("%d/%m/%Y")}</p>
    </div>
    """, unsafe_allow_html=True)

uploaded_file = st.file_uploader("Sube el archivo de órdenes (.csv, .xlsx, .xls)", type=["csv", "xlsx", "xls"])

if uploaded_file is not None:
    try:
        # 1. LECTURA FLEXIBLE
        if uploaded_file.name.endswith('.csv'):
            try:
                df = pd.read_csv(uploaded_file, sep=',', engine='python', encoding='utf-8')
            except:
                df = pd.read_csv(uploaded_file, sep=';', engine='python', encoding='latin-1')
        else:
            df = pd.read_excel(uploaded_file)

        df.columns = df.columns.str.strip()
        all_cols = df.columns.tolist()

        # 2. MAPEO DE COLUMNAS (Sidebar para flexibilidad total)
        st.sidebar.header("⚙️ Configuración de Columnas")
        def detectar(targets):
            for t in targets:
                for col in all_cols:
                    if t.lower() in col.lower(): return all_cols.index(col)
            return 0

        c_fecha = st.sidebar.selectbox("Fecha", all_cols, index=detectar(['Fecha']))
        c_tech = st.sidebar.selectbox("Técnico", all_cols, index=detectar(['Técnico']))
        c_estado = st.sidebar.selectbox("Estado", all_cols, index=detectar(['Estado']))
        c_rep = st.sidebar.selectbox("Repuestos", all_cols, index=detectar(['Repuestos']))
        c_orden = st.sidebar.selectbox("Orden #", all_cols, index=detectar(['Orden', '#']))
        c_prod = st.sidebar.selectbox("Producto", all_cols, index=detectar(['Producto', 'Articulo']))
        c_serie = st.sidebar.selectbox("Serie/Artículo", all_cols, index=detectar(['Serie', 'Cod']))

        # 3. PROCESAMIENTO DE DATOS
        df['Fecha_DT'] = pd.to_datetime(df[c_fecha], dayfirst=True, errors='coerce')
        df['Dias_Antiguedad'] = (hoy - df['Fecha_DT']).dt.days
        df['Alerta'] = df['Dias_Antiguedad'].apply(lambda x: "🚩 CRÍTICO (+15d)" if x > 15 else "OK")

        # --- LÓGICA DE FILTRADO ---
        df['es_go'] = df[c_tech].str.upper().str.startswith('GO', na=False)
        
        # Filtro para Nacionales: Solo estados específicos con repuestos
        mask_solicita = df[c_estado].str.contains('Solicita', case=False, na=False)
        mask_proceso = df[c_estado].str.contains('Proceso/Repuestos', case=False, na=False)
        mask_tiene_rep = df[c_rep].astype(str).str.strip().apply(lambda x: len(x) > 2 and x.lower() != 'nan')
        
        # Filtro Final: Todo lo que sea GO + Nacionales con repuestos pendientes
        df_filtrado = df[ (df['es_go'] == True) | (mask_solicita | (mask_proceso & mask_tiene_rep)) ].copy()

        # Ordenar: GO primero (Prioridad 0), resto (Prioridad 1)
        df_filtrado['Prioridad'] = df_filtrado['es_go'].map({True: 0, False: 1})
        df_filtrado = df_filtrado.sort_values(by=['Prioridad', c_tech, 'Dias_Antiguedad'], ascending=[True, True, False])

        if not df_filtrado.empty:
            # MÉTRICAS
            m1, m2, m3 = st.columns(3)
            m1.metric("📦 Órdenes Totales", len(df_filtrado))
            m2.metric("🏢 Centros GO", len(df_filtrado[df_filtrado['es_go']]))
            m3.metric("🚩 Críticas (>15d)", len(df_filtrado[df_filtrado['Dias_Antiguedad'] > 15]))

            # PREPARACIÓN EXCEL
            df_filtrado['Enviado'] = "[  ]"
            # Se agregan Producto y Serie/Artículo a la lista final
            cols_finales = ['Enviado', 'Alerta', c_orden, c_fecha, c_tech, c_estado, c_prod, c_serie, c_rep, 'Dias_Antiguedad']
            
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_filtrado[cols_finales].to_excel(writer, index=False, sheet_name='Reporte')
                ws = writer.sheets['Reporte']
                
                if EXCEL_STYLING:
                    header_fill = PatternFill(start_color='1F4E78', end_color='1F4E78', fill_type='solid')
                    go_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
                    sep_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')

                    # Cabeceras
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = Font(color='FFFFFF', bold=True)
                        cell.alignment = Alignment(horizontal='center')

                    # Estilo por filas
                    for r in range(2, ws.max_row + 1):
                        tech_val = str(ws.cell(row=r, column=cols_finales.index(c_tech)+1).value)
                        if tech_val.upper().startswith('GO'):
                            for c in range(1, len(cols_finales)+1):
                                ws.cell(row=r, column=c).fill = go_fill

                    # Separadores de Técnico
                    row = 2
                    while row <= ws.max_row:
                        idx_t = cols_finales.index(c_tech) + 1
                        curr = ws.cell(row=row, column=idx_t).value
                        prev = ws.cell(row=row-1, column=idx_t).value if row > 2 else None
                        if prev and curr != prev and prev != c_tech:
                            ws.insert_rows(row)
                            ws.cell(row=row, column=1).value = f"📍 SECCIÓN: {curr}"
                            ws.cell(row=row, column=1).fill = sep_fill
                            ws.cell(row=row, column=1).font = Font(bold=True)
                            row += 1
                        row += 1

            st.download_button(
                label="📥 Descargar Reporte Completo (GO + Nacional)",
                data=output.getvalue(),
                file_name=f"Reporte_Repuestos_Full_{hoy.strftime('%d-%m')}.xlsx",
                use_container_width=True
            )

            # VISTA WEB
            st.markdown("### Detalle Operativo")
            for taller in df_filtrado[c_tech].unique():
                sub = df_filtrado[df_filtrado[c_tech] == taller]
                es_go = taller.upper().startswith('GO')
                retrasos = len(sub[sub['Dias_Antiguedad'] > 15])
                
                title = f"{'🏢' if es_go else '🔧'} {taller} ({len(sub)} órdenes)"
                if retrasos > 0: title += f" | 🚩 {retrasos} RETRASADAS"
                
                with st.expander(title):
                    st.dataframe(sub[cols_finales], hide_index=True, use_container_width=True)
        else:
            st.warning("No se encontraron órdenes con repuestos pendientes.")

    except Exception as e:
        st.error(f"Error crítico: {e}")
